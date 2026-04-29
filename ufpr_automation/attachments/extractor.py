"""Extract readable text from email attachments.

Supports PDF (PyMuPDF), DOCX (python-docx), XLSX (openpyxl), and plain text.
Images and scanned PDFs are processed with OCR (Tesseract) when available;
if Tesseract is not installed, they are flagged with needs_ocr=True.
"""

from __future__ import annotations

from pathlib import Path

from ufpr_automation.core.models import AttachmentData
from ufpr_automation.utils.logging import logger

# Minimum chars per page to consider a PDF as having extractable text.
# Below this threshold, the PDF is likely scanned/image-based.
_MIN_CHARS_PER_PAGE = 20


def extract_text_from_attachment(att: AttachmentData) -> str:
    """Extract text from an attachment based on its MIME type.

    Updates att.extracted_text and att.needs_ocr in place.

    Returns:
        The extracted text (also stored in att.extracted_text).
    """
    mime = att.mime_type.lower()
    path = Path(att.local_path)

    if not path.exists():
        logger.warning("Anexo nao encontrado: %s", att.local_path)
        return ""

    try:
        if mime == "application/pdf":
            text = _extract_pdf(path)
            # If PDF text extraction failed (scanned), try OCR
            if not text.strip():
                text = _ocr_pdf_scanned(path)
        elif mime in (
            "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            "application/msword",
        ):
            text = _extract_docx(path)
        elif mime in (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "application/vnd.ms-excel",
        ):
            text = _extract_xlsx(path)
        elif mime.startswith("text/"):
            text = path.read_text(encoding="utf-8", errors="replace")
        elif mime.startswith("image/"):
            text = _ocr_image(path)
        else:
            logger.debug("Tipo MIME nao suportado para extracao: %s (%s)", mime, att.filename)
            att.extracted_text = ""
            return ""
    except Exception as e:
        logger.warning("Falha ao extrair texto de '%s': %s", att.filename, e)
        att.needs_ocr = True
        att.extracted_text = ""
        return ""

    if not text.strip():
        att.needs_ocr = True
        att.extracted_text = ""
        return ""

    att.needs_ocr = False
    att.extracted_text = text
    return text


def _extract_pdf(path: Path) -> str:
    """Extract text from a PDF using PyMuPDF.

    Returns empty string on parse failure (e.g. ``code=7: cycle in resources``)
    so the caller falls through to ``_ocr_pdf_scanned`` instead of giving up.
    """
    import pymupdf

    try:
        doc = pymupdf.open(str(path))
    except Exception as e:
        logger.warning("PyMuPDF nao conseguiu abrir '%s' (%s) — tentando OCR", path.name, e)
        return ""

    try:
        pages = []
        for page in doc:
            try:
                text = page.get_text()
            except Exception as e:
                logger.debug("PyMuPDF falhou em pagina de '%s': %s", path.name, e)
                continue
            if text.strip():
                pages.append(text)
        num_pages = len(doc)
    finally:
        doc.close()

    full_text = "\n\n".join(pages)

    # Detect scanned PDFs: very little text relative to page count
    if num_pages > 0 and len(full_text) / num_pages < _MIN_CHARS_PER_PAGE:
        return ""  # Caller will set needs_ocr

    return full_text


def _extract_docx(path: Path) -> str:
    """Extract text from a DOCX file using python-docx."""
    try:
        from docx import Document
    except ImportError:
        logger.warning("python-docx nao instalado — pip install python-docx")
        return ""

    doc = Document(str(path))
    paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
    return "\n".join(paragraphs)


def _extract_xlsx(path: Path) -> str:
    """Extract text from an XLSX file using openpyxl."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        logger.warning("openpyxl nao instalado — pip install openpyxl")
        return ""

    wb = load_workbook(str(path), read_only=True, data_only=True)
    lines = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        lines.append(f"[Planilha: {sheet}]")
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            line = " | ".join(cells)
            if line.strip(" |"):
                lines.append(line)
    wb.close()
    return "\n".join(lines)


# ---------------------------------------------------------------------------
# OCR — Tesseract-based text extraction for images and scanned PDFs
# ---------------------------------------------------------------------------


def _configure_tesseract():
    """Set Tesseract path on Windows if not already on PATH."""
    import sys

    if sys.platform == "win32":
        import shutil

        if not shutil.which("tesseract"):
            import pytesseract

            win_path = r"C:\Program Files\Tesseract-OCR\tesseract.exe"
            from pathlib import Path

            if Path(win_path).exists():
                pytesseract.pytesseract.tesseract_cmd = win_path


def _is_tesseract_available() -> bool:
    """Check if Tesseract OCR is installed and accessible."""
    try:
        import pytesseract

        _configure_tesseract()
        pytesseract.get_tesseract_version()
        return True
    except Exception:
        return False


def _ocr_image(path: Path) -> str:
    """Extract text from an image file using Tesseract OCR.

    Falls back to empty string (with needs_ocr=True) if Tesseract is not installed.
    """
    try:
        import pytesseract
        from PIL import Image
    except ImportError:
        logger.debug(
            "pytesseract/Pillow nao instalado — pip install pytesseract Pillow "
            "(+ instalar Tesseract OS package)"
        )
        return ""

    if not _is_tesseract_available():
        logger.debug("Tesseract nao encontrado no sistema — OCR indisponivel para '%s'", path.name)
        return ""

    try:
        img = Image.open(path)
        text = pytesseract.image_to_string(img, lang="por+eng")
        if text.strip():
            logger.debug("OCR extraiu %d chars de imagem '%s'", len(text.strip()), path.name)
        return text.strip()
    except Exception as e:
        logger.warning("OCR falhou para imagem '%s': %s", path.name, e)
        return ""


def _ocr_pdf_scanned(path: Path) -> str:
    """Extract text from a scanned PDF by converting pages to images and running OCR.

    Uses PyMuPDF to render pages as pixmaps, then Tesseract for OCR.
    Falls back to empty string if Tesseract is not installed.
    """
    try:
        import pytesseract
        from PIL import Image
    except ImportError:
        logger.debug("pytesseract/Pillow nao instalado — pip install pytesseract Pillow")
        return ""

    if not _is_tesseract_available():
        logger.debug("Tesseract nao encontrado — OCR indisponivel para PDF '%s'", path.name)
        return ""

    try:
        import io

        import pymupdf

        doc = pymupdf.open(str(path))
        pages_text = []

        for page_num, page in enumerate(doc):
            # Render page to image at 300 DPI for good OCR quality
            pix = page.get_pixmap(dpi=300)
            img_data = pix.tobytes("png")
            img = Image.open(io.BytesIO(img_data))

            text = pytesseract.image_to_string(img, lang="por+eng")
            if text.strip():
                pages_text.append(text.strip())

        doc.close()

        full_text = "\n\n".join(pages_text)
        if full_text:
            logger.debug(
                "OCR extraiu %d chars de PDF escaneado '%s' (%d paginas)",
                len(full_text),
                path.name,
                len(pages_text),
            )
        return full_text

    except Exception as e:
        logger.warning("OCR falhou para PDF '%s': %s", path.name, e)
        return ""
